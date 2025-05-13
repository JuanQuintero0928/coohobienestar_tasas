from django.shortcuts import render
from django.views.generic import ListView, TemplateView
from django.http import HttpResponse
from django.db.models import Sum

from historico.models import HistorialPagos
from parametro.models import FormaPago
from funciones.function import separarFecha
from datetime import datetime

#Libreria para generar el Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class InformacionReporte(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/informacion.html'
        return render(request, template_name)

class VerPagosFecha(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/pagosPorFecha.html'
        return render(request, template_name)
    
    def post(self, request, *args, **kwargs):
        fechaInicialForm = request.POST['fechaInicial']
        fechaFinalForm = request.POST['fechaFinal']
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        fechaInicial = separarFecha(fechaInicialForm, 'inicial')
        fechaFinal = separarFecha(fechaFinalForm, 'final')
        # consulta por rango de fecha inicial y final
        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal]
        )
        template_name = 'reporte/pagosPorFecha.html'
        return render(request, template_name, {'query':queryHistorial,'post':'yes', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm})

class ReporteExcelPago(TemplateView):
    def get(self, request, *args, **kwargs):
        fechaInicialForm = request.GET['fechaInicial']
        fechaFinalForm = request.GET['fechaFinal']
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario, formato Y-m-d 00:00:00-05:00
        fechaInicial = separarFecha(fechaInicialForm, 'inicial')
        fechaFinal = separarFecha(fechaFinalForm, 'final')
        # consulta por rango de fecha inicial y final
        queryPagos = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal]
        )

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")  # Fuente en negrita, tamaño 12 y color blanco
        bold_font2 = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 12 y color blanco
        alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación al centro
        fill = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")  # Relleno verde sólido

        # Convertir la cadena de fecha a un objeto datetime
        fecha_objeto1 = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
        # Formatear la fecha en el formato d-m-Y
        fecha_formateada1 = fecha_objeto1.strftime("%d-%m-%Y")
        # Convertir la cadena de fecha a un objeto datetime
        fecha_objeto2 = datetime.strptime(fechaFinalForm, "%Y-%m-%d")
        # Formatear la fecha en el formato d-m-Y
        fecha_formateada2 = fecha_objeto2.strftime("%d-%m-%Y")

        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws.title = 'Pagos'
        titulo1 = f"Reporte de Pagos desde {fecha_formateada1} al {fecha_formateada2}"
        ws['A1'] = titulo1    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:L1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = alignment_center
        ws['A1'].fill = fill

        ws['A2'] = 'Número Registro'
        ws['B2'] = 'Unidad de Servicio'
        ws['C2'] = 'Número Documento'
        ws['D2'] = 'Primer Nombre'
        ws['E2'] = 'Segundo Nombre'
        ws['F2'] = 'Primer Apellido'
        ws['G2'] = 'Segundo Apellido'
        ws['H2'] = 'Mes Pago'
        ws['I2'] = 'Valor Pago'
        ws['J2'] = 'Diferencia'
        ws['K2'] = 'Fecha Pago'
        ws['l2'] = 'Forma Pago'
        
        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,13):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 19
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 11
        ws.column_dimensions['J'].width = 11
        ws.column_dimensions['K'].width = 13
        ws.column_dimensions['L'].width = 13        

        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        for pago in queryPagos:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 1).value = i
            ws.cell(row = cont, column = 2).value = pago.asociado.hogarinfantil             
            ws.cell(row = cont, column = 3).value = int(pago.asociado.numDocumento)
            ws.cell(row = cont, column = 4).value = pago.asociado.nombre1
            ws.cell(row = cont, column = 5).value = pago.asociado.nombre2
            ws.cell(row = cont, column = 6).value = pago.asociado.apellido1
            ws.cell(row = cont, column = 7).value = pago.asociado.apellido2
            ws.cell(row = cont, column = 8).value = pago.mesPago.concepto
            ws.cell(row = cont, column = 9).value = pago.valorPago
            ws.cell(row = cont, column = 10).value = pago.diferencia
            ws.cell(row = cont, column = 11).value = pago.fechaPago.strftime("%d/%m/%Y")
            ws.cell(row = cont, column = 12).value = pago.formaPago.formaPago
            i+=1
            cont+=1

        nombre_archivo = f"Reporte Pago_{fechaInicialForm}-{fechaFinalForm}.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class VerConciliacionBancaria(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/listarConciliacionBancaria.html'
        formaPago = FormaPago.objects.all().order_by('formaPago')
        return render(request, template_name, {'formaPago': formaPago})
    
    def post(self, request, *args, **kwargs):
        formaPago = FormaPago.objects.all().order_by('formaPago')
        fechaInicialForm = request.POST['fechaInicial']
        fechaFinalForm = request.POST['fechaFinal']
        banco = int(request.POST['banco'])
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        fechaInicial = separarFecha(fechaInicialForm, 'inicial')
        fechaFinal = separarFecha(fechaFinalForm, 'final')
        # consulta por rango de fecha inicial y final
        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal],
            formaPago_id=banco
            ).values(
                'asociado__id',
                'asociado__nombre1',
                'asociado__nombre2',
                'asociado__apellido1',
                'asociado__apellido2',
                'asociado__numDocumento',
                'formaPago_id__formaPago',
                'fechaPago',
                'asociado__hogarinfantil',
            ).annotate(total_pagado=Sum('valorPago')).order_by('fechaPago')
        template_name = 'reporte/listarConciliacionBancaria.html'
        return render(request, template_name, {'query':queryHistorial,'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm, 'formaPago':formaPago, 'banco':banco})

class ExcelConciliacionBancaria(TemplateView):
    def get(self, request, *args, **kwargs):
        fechaInicialForm = request.GET['fechaInicial']
        fechaFinalForm = request.GET['fechaFinal']
        banco = int(request.GET['banco'])
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        fechaInicial = separarFecha(fechaInicialForm, 'inicial')
        fechaFinal = separarFecha(fechaFinalForm, 'final')
        # consulta por rango de fecha inicial y final
        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal],
            formaPago_id=banco
            ).values(
                'asociado__id',
                'asociado__nombre1',
                'asociado__nombre2',
                'asociado__apellido1',
                'asociado__apellido2',
                'asociado__numDocumento',
                'formaPago_id__formaPago',
                'fechaPago',
                'asociado__hogarinfantil',
            ).annotate(total_pagado=Sum('valorPago')).order_by('fechaPago')

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")  # Fuente en negrita, tamaño 12 y color blanco
        bold_font2 = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 12 y color negro
        alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación al centro
        fill = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")  # Relleno verde sólido

        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws.title = 'Conciliación Bancaria'
        titulo1 = f"Conciliación Bancaria"
        ws['A1'] = titulo1    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:J1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = alignment_center
        ws['A1'].fill = fill

        ws['A2'] = 'Número Registro'
        ws['B2'] = 'Número Documento'
        ws['C2'] = 'Primer Nombre'
        ws['D2'] = 'Segundo Nombre'
        ws['E2'] = 'Primer Apellido'
        ws['F2'] = 'Segundo Apellido'
        ws['G2'] = 'Movimiento'
        ws['H2'] = 'Valor'
        ws['I2'] = 'Fecha Movimiento'
        ws['J2'] = 'Unidad de Servicio'


        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,11):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14

        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        
        for query in queryHistorial:

            fecha_pago = query['fechaPago']
            fecha_formateada = fecha_pago.strftime("%d/%m/%Y") if fecha_pago else ""
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 1).value = i                   
            ws.cell(row = cont, column = 2).value = int(query['asociado__numDocumento'])
            ws.cell(row = cont, column = 3).value = query['asociado__nombre1']
            ws.cell(row = cont, column = 4).value = query['asociado__nombre2']
            ws.cell(row = cont, column = 5).value = query['asociado__apellido1']
            ws.cell(row = cont, column = 6).value = query['asociado__apellido2']
            ws.cell(row = cont, column = 7).value = query['formaPago_id__formaPago']
            ws.cell(row = cont, column = 8).value = query['total_pagado']
            ws.cell(row = cont, column = 9).value = fecha_formateada
            ws.cell(row = cont, column = 10).value = query['asociado__hogarinfantil']
        
            i+=1
            cont+=1

        nombre_archivo = f"Reporte_Conciliacion_Bancaria.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
